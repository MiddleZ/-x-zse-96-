import re
import subprocess
import hashlib
import requests
import json
from tqdm import tqdm
import time
import pandas as pd
from openpyxl import load_workbook

#代码中所有的 相关网址均已进行 脱敏处理，相关使用者自行负责

url_token = 'ponyma'#待爬取用户的url_token
data_path = ''

def user_followers():
    #读取页数
    with open("pageN.txt",'r',encoding='utf-8') as fp:
        page_N = int(fp.read())
    print('正在读取'+str(page_N)+'页')
    #拼接
    offset = (page_N-1) * 20
    ta = "101_3_3.0"
    cookie_dc0 = "\"APDcLdkxgxOPTiJpWgSNn6GikHyNF0VMK3E=|1628001843\""
    uh = "/api/v4/members/"+url_token+"/followers?include=data%5B*%5D.answer_count%2Carticles_count%2Cgender%2Cfollower_count%2Cis_followed%2Cis_following%2Cbadge%5B%3F%28type%3Dbest_answerer%29%5D.topics&offset="+str(offset)+"&limit=20"
    tf = ta+'+'+uh+'+'+cookie_dc0
    print(tf)
    #md5 加密
    md5 = hashlib.md5()
    md5.update(tf.encode('utf-8'))
    tt = md5.hexdigest() #md5 加密后
    print(tt)
    #修改js文件
    location_herf = 'https://www.脱敏处理.com/people/'+url_token+'/followers?page='+str(page_N) #
    with open("jzq.js",'r',encoding="gbk") as jsp:
        js_code = jsp.read()
    pattern1 = re.compile(r"var pageN = '(.*?)'")
    pattern2 = re.compile(r"console.log\(window.l\(1514\)\[\'ZP\']\('(.*?)'\)\);")
    match1 = pattern1.search(js_code)
    if match1:
        js_code = js_code.replace(match1.group(1),location_herf)
    match2 = pattern2.search(js_code)
    if match2:
        js_code = js_code.replace(match2.group(1),tt)
    with open("jzq.js",'w',encoding='gbk') as jsp:
        jsp.write(js_code)
    #计算x-zse-96参数
    res = subprocess.run("node jzq.js",stdout=subprocess.PIPE,shell=True)
    x = []
    ps = str(res.stdout).split('\\n')
    for p in ps:
        x.append(p)
    print(x[-2])
    t0 = x[-2]
    x_96 = '2.0_' + t0
    #爬虫准备
    url = "https://www.脱敏处理.com/api/v4/members/"+url_token+"/followers?include=data%5B*%5D.answer_count%2Carticles_count%2Cgender%2Cfollower_count%2Cis_followed%2Cis_following%2Cbadge%5B%3F%28type%3Dbest_answerer%29%5D.topics&offset="+str(offset)+"&limit=20"
    params = {
            'include': 'data[*].is_normal,admin_closed_comment,reward_info,is_collapsed,annotation_action,annotation_detail,collapse_reason,is_sticky,collapsed_by,suggest_edit,comment_count,can_comment,content,editable_content,attachment,voteup_count,reshipment_settings,comment_permission,created_time,updated_time,review_info,relevant_info,question,excerpt,is_labeled,paid_info,paid_info_content,reaction_instruction,relationship.is_authorized,is_author,voting,is_thanked,is_nothelp,is_recognized;data[*].mark_infos[*].url;data[*].author.follower_count,vip_info,badge[*].topics;data[*].settings.table_of_content.enabled',
            # 'offset': '0',
            # 'limit': '5',
            'sort_by': 'default',
            'platform': 'desktop',
        }
    headers = {
        'accept':'*/*',
        'Accept-Language':'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
       'Accept-Encoding':'gzip, deflate, br',
        'X-Requested-With':'fetch',
        'Referer':'https://www.脱敏处理.com/people/'+url_token+'/followers?page='+str(page_N),
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
        "cookie":'d_c0="APDcLdkxgxOPTiJpWgSNn6GikHyNF0VMK3E=|1628001843"; _xsrf=HWUgSh20t0DqcQLELvW9aiBDUF9b33tF; q_c1=6b8c51f29aec40b3b5de4a5680882bf3|1645958464000|1645958464000; _zap=b0db2c52-aac3-45a9-a4b5-63fc4e67e0a7; z_c0=2|1:0|10:1697629107|4:z_c0|80:MS4xTF91WkNBQUFBQUFtQUFBQVlBSlZUYXVVRjJZaWVpSXRndFd6RWdCUmwyY2VPdUh1M0djcnZRPT0=|3d38eb60e492abff77f7753d24cb396b164247ad6275aead6368921d317f57af; tst=r; SESSIONID=nBWT56da7Tfbt9dMoNPWtyCxaWHtbdSNAKJ9HiQFaPI; KLBRSID=2177cbf908056c6654e972f5ddc96dc2|1698741524|1698739803',
        "x-zse-93": "101_3_3.0",
        'x-zse-96':x_96
    }
    #爬虫开始
    responses = requests.get(url=url,headers=headers)
    resp_code = responses.status_code
    if resp_code == 200:
        print('状态码'+str(resp_code))

    print(responses.json())

    #json_str = responses.json()
    data = responses.json()
    #解析json串
    #data = json.loads(responses.json())
    data_list = data['data']

    df = pd.DataFrame(data_list, columns=['id', 'url_token','use_default_avatar', 'avatar_url','name', 'headline', 'gender'])
    #数据存储
    book = load_workbook('output.xlsx')

    writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')
    writer.book = book

    writer.sheets = {ws.title:ws for ws in book.worksheets}

    if 'Sheet1' not in writer.sheets:
        df.to_excel(writer, index=False, header=True)
    else:
        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

    writer.save()
    writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')
    df.to_excel(writer, index=False)
    writer.close()
    with open('pageN.txt','w',encoding='utf-8') as fp:
        fp.write(str(page_N+1))

if __name__ == "__main__":
    with open('pageN.txt','r',encoding='utf-8') as pagen:
         xs = int(pagen.read())
    for i in range(9000):
        user_followers()
        for j in tqdm(range(1, 101)):
            time.sleep(0.1)